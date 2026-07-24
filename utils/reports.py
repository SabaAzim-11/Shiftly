"""ShiftSync AI – Report generation
PDF via ReportLab, Excel via OpenPyXL
"""

from __future__ import annotations
import io
from datetime import date
import pandas as pd


def generate_excel_report(df: pd.DataFrame, sheet_name: str = "Report",
                           title: str = "ShiftSync Report") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    ws.merge_cells(f"A1:{chr(65+len(df.columns)-1)}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1A2D5A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{chr(65+len(df.columns)-1)}2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Generated: {date.today().strftime('%d %B %Y')} | ShiftSync AI Platform"
    sub_cell.font = Font(size=10, color="9AA3C0")
    sub_cell.fill = PatternFill("solid", fgColor="161B27")
    sub_cell.alignment = Alignment(horizontal="center")

    header_fill = PatternFill("solid", fgColor="252D3F")
    header_font = Font(bold=True, color="E8EAF0", size=11)
    border = Border(
        bottom=Side(style="thin", color="2A3350"),
        right=Side(style="thin", color="2A3350"),
    )
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=str(col_name).title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[3].height = 22

    row_fills = [PatternFill("solid", fgColor="161B27"), PatternFill("solid", fgColor="1E2435")]
    data_font = Font(color="9AA3C0", size=10)
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 4):
        fill = row_fills[row_idx % 2]
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() if len(df) else 0)
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 35)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def attendance_excel(df: pd.DataFrame) -> bytes:
    return generate_excel_report(df, "Attendance", "ShiftSync – Attendance Report")


def employee_excel(df: pd.DataFrame) -> bytes:
    return generate_excel_report(df, "Employees", "ShiftSync – Employee Directory")


def leave_excel(df: pd.DataFrame) -> bytes:
    return generate_excel_report(df, "Leaves", "ShiftSync – Leave Report")


def generate_pdf_report(df: pd.DataFrame, title: str = "ShiftSync Report") -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                         Paragraph, Spacer, HRFlowable)
        from reportlab.lib.colors import HexColor
    except ImportError:
        return b"PDF generation requires reportlab. Install: pip install reportlab"

    buf = io.BytesIO()
    page_size = landscape(A4) if len(df.columns) > 6 else A4
    doc = SimpleDocTemplate(buf, pagesize=page_size,
                             leftMargin=1.5*cm, rightMargin=1.5*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)

    dark_bg   = HexColor("#0F1117")
    card_bg   = HexColor("#161B27")
    header_bg = HexColor("#1A2D5A")
    accent    = HexColor("#4F8EF7")
    text_c    = HexColor("#E8EAF0")
    text2_c   = HexColor("#9AA3C0")
    border_c  = HexColor("#2A3350")
    alt_row   = HexColor("#1E2435")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        textColor=text_c, fontSize=18, spaceAfter=4,
        fontName="Helvetica-Bold"
    )
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"],
        textColor=text2_c, fontSize=9, spaceAfter=12,
    )

    elements = []
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(
        f"Generated: {date.today().strftime('%d %B %Y')} | ShiftSync AI Platform",
        sub_style
    ))
    elements.append(HRFlowable(width="100%", thickness=1, color=border_c, spaceAfter=10))

    headers = [str(c).title() for c in df.columns]
    data = [headers] + df.astype(str).values.tolist()

    col_count = len(df.columns)
    available_w = (page_size[0] - 3*cm)
    col_w = available_w / col_count

    tbl = Table(data, colWidths=[col_w]*col_count, repeatRows=1)
    tbl_style = TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), header_bg),
        ("TEXTCOLOR",   (0, 0), (-1, 0), text_c),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 9),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [card_bg, alt_row]),
        ("TEXTCOLOR",   (0, 1), (-1, -1), text2_c),
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("GRID",        (0, 0), (-1, -1), 0.5, border_c),
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])
    tbl.setStyle(tbl_style)
    elements.append(tbl)

    doc.build(elements)
    return buf.getvalue()
